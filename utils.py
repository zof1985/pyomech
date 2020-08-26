


# IMPORTS


import numpy as np
import pandas as pd
import openpyxl as xl
import time
import os



# METHODS



def polynome_txt(fit, magnitude_threshold=2, digits=3):
    """
    convert a fitting polynome to its mathematical string

    Input:
        fit: (ndarray, list)
            a list of float values defining the coefficients.

        magnitude_threshold: (float)
            the threshold that has not to be passed in order to avoid the use of the scientific notation
            for the representation of each coefficient.

        digits: (int)
            the decimal precision to be used for each coefficient

    Output:
        txt: (str)
            a string in the form: Y = aX^n + bX^n-1 ....
    """

    # ensure fit is a 1D array
    fit = np.atleast_1d(fit)
    assert fit.ndim == 1, "'fit' must have 1 dimension."

    # get the formatter
    formatter = "{:0." + str(digits) + "f}"

    # get the output string
    txt = r"$Y="
    for i, v in enumerate(fit):

        # add the sign
        sign = "+" if v >= 0 else "-"
        v = abs(v)

        # add the coefficient
        mag =  magnitude(float(v))
        if mag > 10 ** magnitude_threshold or v < 10 ** -magnitude_threshold:
            v_txt = formatter.format(v / (10 ** mag)) + "\cdot" + str(10) + "^{" + str(mag) + "}"
        else:
            v_txt = formatter.format(v)

        # agg the X in the correct order
        order = len(fit) - i - 1
        x_txt = ""
        if order > 0:
            x_txt += "\ X"
        if order > 1:
            x_txt += "^{" + str(order) + "}"

        # aggregate all
        txt += sign + v_txt + x_txt
    txt += "$"

    # return the text
    return txt



def classcheck(obj, classes, txt=None):
	"""
	Check if obj is any of the classes provided. If not an error is thrown with message txt.

	Input:
		obj: (object)
			the obkect to be tested.
		classes: (list)
			a list of str reflecting the names of the classes to be tested.
		txt: (None or str)
			the error message to be thrown in case of unmatch of the obj class with classes. If None, an automatic
			message is thrown.
	"""

	# check classes
	assert classes.__class__.__name__ in ["list", "ndarray"], "'classes' must be a list."

	# check txt
	if txt is None: txt = "'obj' must be any of " + str(classes) + "."
	assert txt.__class__.__name__ == "str", "'txt' must be a str."

	# perform the test
	assert any([obj.__class__.__name__[:len(i)] == i for i in classes]), txt



def magnitude(value, base=10):
    """
    return the order in the given base of the value

    Input
        value: (float)
            the value to be checked
        base: (float)
            the base to be used to define the order of the number

    Output
        mag: (float)
            the number required to elevate the base to get the value
    """

	# return the magnitude
    if value == 0 or base == 0:
        return 0
    else:
        return int(np.floor(np.log(abs(value)) / np.log(base)))



def get_files(path, extension=None, check_subfolders=False):
    """
    list all the files having the required extension in the provided folder and its subfolders (if required).

    Input:
        path: (str)
            a directory where to look for the files.
        extension: (str)
            a str object defining the ending of the files that have to be listed.
        check_subfolders: (bool)
            if True, also the subfolders found in path are searched, otherwise only path is checked.

    Output:
        files: (list)
            a list containing the full_path to all the files corresponding to the input criteria.
    """

    # output storer
    out = []

    # surf the path by the os. walk function
    for root, dirs, files in os.walk(path):
        out += [os.path.join(root, obj) for obj in files if obj[-len(extension):] == extension]
        
        # handle the subfolders
        if not check_subfolders:
            break 

    # return the output
    return out
    


def lvlup(file):
	"""
	Goes to the superior level in the directory path.

	Input:
		file: (str)
			a file or a directory. Otherwise a message is casted and the function returns the input as is

	Output:
		a string reflecting the superior directory of file.
	"""

	# check whether file exists as "file" or directory
	return os.path.abspath(os.path.join(file, os.pardir))



def to_excel(P, D, N="Sheet1", keep_index=True, new_file=False):
	"""
	a shorthand function to save a pandas dataframe to an excel file

	Input
		
		P: (str)
			the path to the file where to store the file.

		D: (DataFrame)
			a dataframe.

		keep_index: (boolean)
			if True, the dataframe index is preserved. Otherwise it is ignored.

		new_file: (boolean)
			if True, a completely new file will be created.

	Output
		The functions stores the data to the indicated file.
	"""

	# get the workbook
	if os.path.exists(P) and not new_file:
		wb = xl.load_workbook(P)
	else:
		wb = xl.Workbook()
		try:
			sh = wb["Sheet"]
			wb.remove(sh)
		except Exception:
			pass

	# get the sheet
	try:
		sh = wb[N]
		wb.remove(sh)
	except Exception:
		pass
	sh = wb.create_sheet(N)

	# write the headers
	[R, C] = D.shape
	if keep_index:
		index = np.atleast_2d(D.index.tolist())
		if index.shape[0] == 1:
			index = index.T
		data_cols = index.shape[1] + 1
	else:
		data_cols = 1
	header = np.atleast_2d(D.columns.tolist())
	if header.shape[0] == 1:
		header = header.T
	data_rows = header.shape[1] + 1
	for i, col in enumerate(header):
		for j, el in enumerate(col):
			ch = data_cols + i
			rh = 1 + j
			sh.cell(rh, ch, el)
	if keep_index:
		for i, row in enumerate(index):
			for j, el in enumerate(row):
				ri = data_rows + i
				ci = 1 + j
				sh.cell(ri, ci, el)

	# write the data
	V = D.values
	[sh.cell(data_rows + r, data_cols + c, V[r, c]) for r in range(R) for c in range(C)]

	# save data
	wb.save(P)



def from_excel(path, sheets=None, verbose=False, **kwargs):
	"""
	a shorthand function to collect data from an excel file and to store them into a dict.

	Input
		path:		(str)
					the path to the file where to store the file.

		sheets:		(list of str)
					the name of the sheets to be imported. If None all sheets will be imported.

		verbose:	(bool)
					if True, messages are generated if errors are encountered.

		kwargs:		additional arguments passed to pandas.read_excel

	Output
		a dict object with keys equal to the sheets name and pandas dataframe as elements of each sheet
		in the excel file.
	"""

	# retrive the data in the path file
	try:
		xlfile = pd.ExcelFile(path)
	except Exception:
		if verbose:
			print(path + " is not valid.")
		try:
			xlfile.close()
		except Exception:
			pass
		return None

	# get the sheets name
	sheets = np.array(xlfile.sheet_names if sheets is None else [sheets]).flatten()

	# get the data
	dfs = {i: pd.read_excel(path, i, **kwargs) for i in sheets}

	# close the excel file
	xlfile.close()

	# return the dict
	return dfs



def get_time(tic=None, toc=None, as_string=True, compact=True):
	"""
	get the days, hours, minutes and seconds between the two times. If only tic is provided, it is considered as
	the lapsed time. If neither tic nor toc are provided, the function returns the current time as float

	Input (optional)
		tic: (int)
			an integer representing the starting time
		toc: (int)
			an integer indicating the stopping time
		as_string: (bool)
			should the output be returned as string?
		compact: (bool)
			if "as_string" is true, should the time be reported in a compact or in an extensive way?

	Output
		If nothing is provided, the function returns the current time. If only tic is provided, the function
		returns the time value from it to now. If both tic and toc are provided, the function returns the time
		difference between them.
	"""

	# check what to do
	if tic is None: return time.time()
	elif toc is None: tm = np.float(tic)
	else: tm = np.float(toc - tic)

	# convert the time value in days, hours, minutes, seconds and milliseconds
	d = int(np.floor(tm / 86400))
	tm -= (d * 86400)
	h = int(np.floor(tm / 3600))
	tm -= (h * 3600)
	m = int(np.floor(tm / 60))
	tm -= (m * 60)
	s = int(np.floor(tm))
	tm -= s
	ms = int(np.round(1000 * tm, 0))

	# report the calculated time
	if not as_string:
		return {"Days": [d], "Hours": [h], "Minutes": [m], "Seconds": [s], "Milliseconds": [ms]}
	else:
		st = "{:0>2d}".format(d) + (" Days - " if not compact else ":")
		st += "{:0>2d}".format(h)
		st += " Hours - " if not compact else ":"
		st += "{:0>2d}".format(m)
		st += " Minutes - " if not compact else ":"
		st += "{:0>2d}".format(s)
		st += " Seconds - " if not compact else ":"
		st += "{:0>3d}".format(ms)
		st += " Milliseconds" if not compact else ""
		return st



def extract(data, factors, axis=1):
	"""
	Extract from data the names according to factors.

	Input:
		data: (DataFrame)
			a dataframe containing the parameters
		factors: (dict)

	Output:
		a list where:
			0. the extracted dataframe having the columns of the factors and of the indicated names
			1. the index of data to corresponding to the extracted values
	"""
	
	# return the wanted rows
	wnt = data[[i for i in factors]].isin(factors).all(axis)

	# get the data and its index
	df = data.loc[wnt]
	ix = df.index.to_numpy()

	# return the subset
	return df, ix
